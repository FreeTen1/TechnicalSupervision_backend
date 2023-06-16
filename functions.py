from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
from sqlalchemy import desc
import sqlalchemy
from general_function import compare_dates, general_filter, hours_between_dates
from models import Artist, Contractor, DayType, PaidStatus, ResponsibleDepartment, StatusesExecution, StatusesK, Supervision, User
from my_engine import session_scope
from io import BytesIO


def authorization(login: str, password: str) -> tuple:
    with session_scope() as session:
        user = session.query(User).filter(User.login == login,
                                          User.password == password).one_or_none()
    
        if not user:
            return {"message": "Неверный логин или пароль"}, 401
        else:        
            return user.as_dict(), 200


def get_lists():
    result = {}
    with session_scope() as session:
        result["artists"] = [item.as_dict() for item in session.query(Artist).all()]
        result["contractors"] = [dict_item for item in session.query(Contractor).all() if (dict_item := item.as_dict()) and dict_item["is_archived"] == 0]
        result["day_types"] = [item.as_dict() for item in session.query(DayType).all()]
        result["paid_statuses"] = [item.as_dict() for item in session.query(PaidStatus).all()]
        result["statuses_execution"] = [item.as_dict() for item in session.query(StatusesExecution).all()]
        result["statuses_ks"] = [item.as_dict() for item in session.query(StatusesK).all()]
        result["responsible_departments"] = [item.as_dict() for item in session.query(ResponsibleDepartment).all()]
    
    return result, 200


def create_new_artist(artist_fio: str) -> int:
    """
    Функция по созданию нового артиста.\n
    В случае если такое fio уже есть в списке, возвращает id найденного исполнителя.\n
    Возвращает id нового или найденного исполнителя
    """
    with session_scope() as session:
        artist: Artist = session.query(Artist).filter(Artist.fio == artist_fio).one_or_none()
        if artist:
            artist_id = artist.id
        else:
            artist: Artist = Artist(fio=artist_fio)
            session.add(artist)
            session.commit()
            artist_id = artist.id

    return artist_id


def add_supervision(kwargs: dict) -> tuple:
    """Добавление нового тех. надзора"""
    kwargs = general_filter(kwargs)
    if kwargs.get("artist"):
        kwargs["artist_id"] = create_new_artist(kwargs["artist"])

    kwargs.pop("artist")

    try:
        with session_scope() as session:
            session.add(Supervision(**kwargs))
    except sqlalchemy.exc.IntegrityError:
        return {"message": "Вы пытаетесь добавить некорректное значение списка"}, 400
    except sqlalchemy.exc.DataError:
        return {"message": "Вы пытаетесь добавить некорректное значение даты"}, 400

    return None, 200


def get_supervisions(my_filter: dict) -> tuple:
    """Получение всех тех. надзоров, с фильтрацией (если нужна) и сортировкой"""
    query = [Supervision.is_archived == 0]
    if my_filter.get("date_start") and my_filter.get("date_end"):
        if compare_dates(my_filter["date_start"], my_filter["date_end"]):
            query.append(Supervision.datetime_start.between(my_filter["date_start"], f'{my_filter["date_end"]} 23:59'))
        else:
            return {"message": "Дата начала должна быть меньше либо равна дате конца"}, 400
    elif my_filter.get("year") and my_filter.get("month"):
        query.append(Supervision.datetime_start.like(f'{my_filter["year"]:04d}-{my_filter["month"]:02d}%'))
    else:
        return {"message": "Не переданы необходимые параметры"}, 400

    if my_filter.get("contractor_id"):
        query.append(Supervision.contractor_id == my_filter["contractor_id"])
    
    if my_filter.get("status_ks_id"):
        query.append(Supervision.status_ks_id == my_filter["status_ks_id"])
    
    if my_filter.get("status_execution_id"):
        query.append(Supervision.status_execution_id == my_filter["status_execution_id"])
    
    order = {"id": Supervision.id, "datetime_start": Supervision.datetime_start,
             "datetime_end": Supervision.datetime_end, "station": Supervision.station}

    with session_scope() as session:
        if my_filter.get("sort_key") and my_filter.get("sort_by"):
            supervision: Supervision = session.query(Supervision).filter(*query).order_by(order[my_filter["sort_key"]]).all(
            ) if my_filter["sort_by"] == "ASC" else session.query(Supervision).filter(*query).order_by(desc(order[my_filter["sort_key"]])).all()
        else:
            print(my_filter)
            supervision: Supervision = session.query(Supervision).filter(*query).all()
        
        result = [item.as_dict() for item in supervision]
    
    return result, 200


def get_single_supervision(supervision_id: int) -> tuple:
    """Получение информации по конкретному тех.надзору"""
    with session_scope() as session:
        supervision: Supervision = session.query(Supervision).get(supervision_id)
        return (supervision.as_dict(), 200) if supervision else ({"message": f"Тех. надзор с id = {supervision_id} не найден"}, 400)


def change_supervision(supervision_id: int, kwargs: dict) -> tuple:
    """Изменение информации конкретного тех. надзора"""
    kwargs = general_filter(kwargs)
    if kwargs.get("artist"):
        kwargs["artist_id"] = create_new_artist(kwargs["artist"])
    else:
        kwargs["artist_id"] = None

    kwargs.pop("artist")

    try:
        with session_scope() as session:
            supervision: Supervision = session.query(Supervision).filter(Supervision.id == supervision_id).update(kwargs, synchronize_session='fetch')
            if not supervision:
                return {"message": "Вы пытаетесь изменить несуществующий тех. надзор"}, 400

    except sqlalchemy.exc.IntegrityError:
        return {"message": "Вы пытаетесь добавить некорректное значение списка"}, 400
    except sqlalchemy.exc.DataError:
        return {"message": "Вы пытаетесь добавить некорректное значение даты"}, 400

    return None, 200


def delete_supervision(supervision_id: int) -> tuple:
    """Удаление (архивирование) конкретного тех. надзора"""

    with session_scope() as session:
        supervision: Supervision = session.query(Supervision).get(supervision_id)
        supervision.is_archived = 1

    return None, 200


def excel_load(supervision_ids: list, load_type: str) -> BytesIO:
    with session_scope() as session:
        supervisions = session.query(Supervision).filter(Supervision.id.in_(supervision_ids)).all()

        wb = openpyxl.Workbook()
        wb.create_sheet(title=f"Выгрузка", index=0)
        sheet = wb[f"Выгрузка"]
        
        # стили для текста
        alignCenter_wrapText = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wrap_text = Alignment(wrap_text=True)
        font14_bold = Font(size=14, bold=True)
        font14 = Font(size=14, bold=False)
        font22 = Font(size=22, bold=True)
        
        sheet['A1'] = 'План работы МТК'
        if load_type == "outside":
            sheet.merge_cells('A1:M1')
            sheet.append(["№ п/п", "Дата, время начала работ", "Дата, время окончания работ",
                          "Время, всего", "Станция (место проведения работ)", "Отдел, Дистанция", "Ф. И. О. осуществляющего технический надзор", 
                          "Вид проводимой работы (объем произведенной работы)", "Название сторонней организации", 
                          "Фамилия, Имя, телефон производителя", "Номер совместного приказа", "Примечание", "Подпись работника"])

            for index, row in enumerate(supervisions, 1):
                row = row.as_dict()
                sheet.append([index, datetime.strptime(row["datetime_start"], '%Y-%m-%d %H:%M').strftime('%d.%m.%Y %H:%M'), 
                              datetime.strptime(row["datetime_end"], '%Y-%m-%d %H:%M').strftime('%d.%m.%Y %H:%M'), 
                              hours_between_dates(row["datetime_start"], row["datetime_end"]),
                              row["station"], row["department_distance"], row["artist"], row["type_work"], 
                              'test', row["manufacturer_info"],
                              row["order_number"], row["note"], ""])
        
        elif load_type == "inside":
            sheet.merge_cells('A1:N1')
            sheet.append(["№ п/п", "Дата, время начала работ", "Дата, время окончания работ",
                          "Время, всего", "Станция (место проведения работ)", "Отдел, Дистанция", "Ф. И. О. осуществляющего технический надзор", 
                          "Вид проводимой работы (объем произведенной работы)", "Название сторонней организации", 
                          "Фамилия, Имя, телефон производителя", "Номер совместного приказа", "Примечание", "Подпись работника", "Ответственное подразделение"])

            for index, row in enumerate(supervisions, 1):
                row = row.as_dict()
                sheet.append([index, datetime.strptime(row["datetime_start"], '%Y-%m-%d %H:%M').strftime('%d.%m.%Y %H:%M'), 
                              datetime.strptime(row["datetime_end"], '%Y-%m-%d %H:%M').strftime('%d.%m.%Y %H:%M'), 
                              hours_between_dates(row["datetime_start"], row["datetime_end"]),
                              row["station"], row["department_distance"], row["artist"], row["type_work"], 
                              'test', row["manufacturer_info"],
                              row["order_number"], row["note"], "", row["department_responsible"]])


        for column in sheet.columns:
            for cell in column:
                cell.font = font14
                cell.alignment = wrap_text
            
            column[1].font = font14_bold
            column[1].alignment = alignCenter_wrapText

        sheet['A1'].alignment = alignCenter_wrapText
        sheet['A1'].font = font22
        sheet.row_dimensions[1].height = 80

        sheet.column_dimensions['A'].width = 17
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 23
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 22
        sheet.column_dimensions['H'].width = 53
        sheet.column_dimensions['I'].width = 31
        sheet.column_dimensions['J'].width = 27
        sheet.column_dimensions['K'].width = 36
        sheet.column_dimensions['L'].width = 60
        sheet.column_dimensions['M'].width = 20
        sheet.column_dimensions['N'].width = 20

        bytes_book = BytesIO()
        wb.save(bytes_book)
        wb.close()
        bytes_book.seek(0)

        return bytes_book


def take_in_ks(take_in_ks_ids: list, not_take_in_ks_ids: list) -> tuple:
    with session_scope() as session:
        if take_in_ks_ids:
            session.query(Supervision).filter(Supervision.id.in_(take_in_ks_ids)).update({"status_ks_id": 1}, synchronize_session='fetch')
        
        if not_take_in_ks_ids:
            session.query(Supervision).filter(Supervision.id.in_(not_take_in_ks_ids)).update({"status_ks_id": 2, "comment": None}, synchronize_session='fetch')
    
    return None, 200


def supervisions_count_info(year: int) -> tuple:
    months = ("01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12")
    with session_scope() as session:
        all_count = session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}%")).count()
        completed = session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}%"), Supervision.status_execution_id == 2).count()
        take_in_ks = session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}%"), Supervision.status_ks_id == 1).count()
        
        count_months = {}
        for month_num in months:
            count_months[month_num] = {
                "all_count": session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}-{month_num}%")).count(),
                "completed": session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}-{month_num}%"), Supervision.status_execution_id == 2).count(),
                "take_in_ks": session.query(Supervision).filter(Supervision.is_archived == 0, Supervision.datetime_start.like(f"{year}-{month_num}%"), Supervision.status_ks_id == 1).count()
            }

    
    return {"all_count": all_count, "completed": completed, "take_in_ks": take_in_ks, "count_months": count_months}, 200
